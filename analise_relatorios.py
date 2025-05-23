import pandas as pd
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether, PageBreak, Image
from reportlab.lib.units import cm
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Usar backend não interativo
import io
from reportlab.lib.enums import TA_CENTER
from config_manager import ConfigManager

def format_hms(value):
    if pd.isnull(value):
        return ''
    if isinstance(value, pd.Timestamp):
        return value.strftime('%H:%M:%S')
    if isinstance(value, timedelta):
        total_seconds = int(value.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f'{hours:02}:{minutes:02}:{seconds:02}'
    return str(value)

def criar_grafico_hora_a_hora(df):
    """
    Cria gráficos de barras para rotas e pedidos expedidos por hora
    
    Args:
        df: DataFrame com colunas 'Hora', 'Rotas Expedidas', 'Pedidos Expedidos'
    
    Returns:
        Uma lista de objetos BytesIO contendo os gráficos
    """
    if df.empty or 'Total' in df['Hora'].values:
        # Remove a linha de totais para o gráfico
        df = df[df['Hora'] != 'Total'].copy()
    
    if df.empty:
        return []
        
    # Convertendo para valores numéricos para ordenação
    df = df.copy()
    
    # Extrair apenas a hora inicial de cada intervalo (ex: "15:00 às 16:00" -> "15:00")
    df['Hora_Ordenacao'] = df['Hora'].apply(lambda x: x.split(' às ')[0] if ' às ' in str(x) else x)
    
    # Ordenar o DataFrame
    df = df.sort_values('Hora_Ordenacao')
    
    # Adicionar colunas de variação percentual entre períodos
    if len(df) > 1:
        df['Var_Rotas'] = df['Rotas Expedidas'].pct_change() * 100
        df['Var_Pedidos'] = df['Pedidos Expedidos'].pct_change() * 100
        # Preencher NaN da primeira linha com 0
        df['Var_Rotas'] = df['Var_Rotas'].fillna(0)
        df['Var_Pedidos'] = df['Var_Pedidos'].fillna(0)
    else:
        df['Var_Rotas'] = 0
        df['Var_Pedidos'] = 0
    
    # Calcular totais acumulados
    df['Rotas_Acumulado'] = df['Rotas Expedidas'].cumsum()
    df['Pedidos_Acumulado'] = df['Pedidos Expedidos'].cumsum()
    
    # Configurar estilo do gráfico para um visual mais moderno
    plt.style.use('seaborn-v0_8-whitegrid')
    
    # Cores da Shopee - nova paleta
    SHOPEE_MAIN = '#EE4D2D'  # Cor principal (vermelho/laranja)
    SHOPEE_BLUE = '#113366'  # Azul escuro
    SHOPEE_YELLOW = '#FFBB00'  # Amarelo
    SHOPEE_BG = '#FFFFFF'  # Fundo branco para maior contraste
    SHOPEE_TEXT = '#222222'
    SHOPEE_LIGHT_MAIN = '#FFCCBC'
    SHOPEE_GRID = '#DDDDDD'
    
    # Lista para armazenar os buffers de imagem
    graficos_buffer = []
    
    # Função para formatar números com separador de milhar
    def format_number(x):
        return f"{x:,.0f}".replace(",", ".")
    
    # Gráfico combinado de rotas e pedidos (mais moderno)
    fig, ax1 = plt.subplots(figsize=(14, 7.5), facecolor=SHOPEE_BG)
    
    # Configurar estilo geral
    ax1.set_facecolor(SHOPEE_BG)
    fig.patch.set_facecolor(SHOPEE_BG)
    
    # Eixo para rotas
    ax1.set_xlabel('Horário', fontsize=13, color=SHOPEE_TEXT, fontweight='bold')
    ax1.set_ylabel('Número de Rotas', fontsize=13, color=SHOPEE_MAIN, fontweight='bold')
    
    # Configurar grid apenas para linhas horizontais e com estilo mais leve
    ax1.grid(axis='y', linestyle='--', alpha=0.4, color=SHOPEE_GRID)
    ax1.grid(axis='x', linestyle='--', alpha=0.3, color=SHOPEE_GRID)
    ax1.spines['top'].set_visible(False)
    ax1.spines['right'].set_visible(False)
    ax1.spines['left'].set_color(SHOPEE_GRID)
    ax1.spines['bottom'].set_color(SHOPEE_GRID)
    
    # Barras com efeito de gradiente e bordas - usando cor mais sólida
    bars1 = ax1.bar(df['Hora'], df['Rotas Expedidas'], color=SHOPEE_MAIN, 
                   alpha=0.9, label='Rotas', edgecolor=SHOPEE_TEXT, linewidth=0.5,
                   zorder=3)
    
    # Adicionar brilho/reflexo nas barras
    for bar in bars1:
        x = bar.get_x()
        width = bar.get_width()
        height = bar.get_height()
        
        # Pequeno realce na parte superior
        ax1.add_patch(plt.Rectangle((x, height*0.9), width, height*0.1, 
                                  color=SHOPEE_LIGHT_MAIN, alpha=0.6,
                                  zorder=4))
    
    ax1.tick_params(axis='y', labelcolor=SHOPEE_MAIN, labelsize=11)
    ax1.tick_params(axis='x', labelsize=11)
    
    # Adicionar valores das barras - posicionando dentro ou acima das barras conforme a altura
    for bar in bars1:
        height = bar.get_height()
        bar_width = bar.get_width()
        bar_x = bar.get_x() + bar_width/2
        
        # Determinar se o valor vai dentro ou acima da barra
        text_color = SHOPEE_MAIN
        y_position = height + 1  # Posição padrão (acima da barra)
        va = 'bottom'  # Alinhamento vertical padrão
        
        # Para barras com altura suficiente, colocar o texto dentro
        if height > 10:  # Altura mínima para texto interno
            text_color = 'white'
            y_position = height/2  # Centro da barra
            va = 'center'  # Centralizado verticalmente
        
        # Adicionar o texto com caixa de fundo apenas se estiver fora da barra
        if text_color == SHOPEE_MAIN:
            ax1.text(bar_x, y_position, f'{format_number(height)}',
                    ha='center', va=va, fontweight='bold', 
                    color=text_color, fontsize=11, zorder=6,
                    bbox=dict(facecolor='white', alpha=0.9, edgecolor=SHOPEE_MAIN, pad=2, boxstyle='round,pad=0.2'))
        else:
            # Texto dentro da barra, sem caixa
            ax1.text(bar_x, y_position, f'{format_number(height)}',
                    ha='center', va=va, fontweight='bold', 
                    color=text_color, fontsize=11, zorder=6)
    
    # Eixo secundário para pedidos
    ax2 = ax1.twinx()
    ax2.set_ylabel('Número de Pedidos', fontsize=13, color=SHOPEE_BLUE, fontweight='bold')
    line = ax2.plot(df['Hora'], df['Pedidos Expedidos'], 'o-', 
                   color=SHOPEE_BLUE, linewidth=3, markersize=10, 
                   label='Pedidos', zorder=4)
    
    # Adicionar área sob a linha de pedidos
    ax2.fill_between(df['Hora'], df['Pedidos Expedidos'], alpha=0.15, color=SHOPEE_BLUE, zorder=2)
    
    ax2.tick_params(axis='y', labelcolor=SHOPEE_BLUE, labelsize=11)
    ax2.spines['right'].set_color(SHOPEE_BLUE)
    ax2.spines['top'].set_visible(False)
    
    # Desativar o grid para o eixo secundário para evitar duplos grids
    ax2.grid(False)
    
    # Nova abordagem para os valores de pedidos - valor e porcentagem juntos, mas com cores distintas
    for i, (hora, valor) in enumerate(zip(df['Hora'], df['Pedidos Expedidos'])):
        y_offset = 35  # Espaçamento vertical padrão
        var_pct = df['Var_Pedidos'].iloc[i] if i > 0 else 0
        var_txt = f"{var_pct:+.1f}%" if i > 0 and var_pct != 0 else ""
        if var_txt:
            var_color = 'crimson' if var_pct < 0 else '#4CAF50'
            arrow = '▲' if var_pct > 0 else '▼'
            # Desenhar valor com caixa azul
            ax2.annotate(
                f'{format_number(valor)}',
                    (hora, valor),
                    xytext=(0, y_offset), 
                    textcoords='offset points',
                ha='right',
                va='center',
                fontsize=11,
                    fontweight='bold',
                    color=SHOPEE_BLUE,
                    zorder=7,
                bbox=dict(facecolor='white', alpha=0.9, edgecolor=SHOPEE_BLUE, pad=2, boxstyle='round,pad=0.3')
            )
            # Desenhar porcentagem ao lado, sem caixa, cor verde/vermelha
            ax2.annotate(
                f'{arrow} {var_txt}',
                           (hora, valor),
                xytext=(10, y_offset),  # 20px à direita do valor
                           textcoords='offset points',
                ha='left',
                           va='center',
                fontsize=11,
                fontweight='bold',
                           color=var_color,
                zorder=7
            )
        else:
            ax2.annotate(
                f'{format_number(valor)}',
                (hora, valor),
                xytext=(0, y_offset),
                textcoords='offset points',
                ha='center',
                va='center',
                fontsize=11,
                           fontweight='bold',
                color=SHOPEE_BLUE,
                zorder=7,
                bbox=dict(facecolor='white', alpha=0.9, edgecolor=SHOPEE_BLUE, pad=2, boxstyle='round,pad=0.3')
            )
    
    # Removendo o título principal para evitar duplicação com o título da tabela
    # plt.suptitle('Rotas e Pedidos Expedidos por Hora', 
    #          fontsize=20, fontweight='bold', color=SHOPEE_TEXT,
    #          y=0.98)
    
    # Adicionar subtítulo com totais
    total_rotas = df['Rotas Expedidas'].sum()
    total_pedidos = df['Pedidos Expedidos'].sum()
    plt.title(f'Total de Rotas: {format_number(total_rotas)} | Total de Pedidos: {format_number(total_pedidos)}', 
             fontsize=13, color=SHOPEE_TEXT, pad=60)
    
    # Combinar legendas dos dois eixos
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    
    # Adicionar uma linha de referência para média de rotas
    if not df.empty:
        media_rotas = df['Rotas Expedidas'].mean()
        ax1.axhline(y=media_rotas, color=SHOPEE_YELLOW, linestyle='--', 
                   linewidth=2.5, label='Média', zorder=2, alpha=0.9)
        
        # Posicionar o texto da média no lado esquerdo para evitar sobreposição
        ax1.text(df['Hora'].iloc[0], media_rotas + 2, 
                f'Média: {format_number(media_rotas)}', 
                fontsize=11, color=SHOPEE_YELLOW, fontweight='bold',
                bbox=dict(facecolor='white', alpha=0.9, edgecolor=SHOPEE_YELLOW, boxstyle='round,pad=0.3'),
                verticalalignment='bottom')
    
    # Legenda com design melhorado - posicionada fora do gráfico
    fig.legend(lines1 + lines2, ['Rotas', 'Média', 'Pedidos'], 
              loc='upper center', bbox_to_anchor=(0.5, 0.94),
              ncol=3, frameon=True, framealpha=0.95,
              facecolor=SHOPEE_BG, edgecolor=SHOPEE_GRID,
              fontsize=12, columnspacing=1.5)
    
    # Ajustar as margens para acomodar apenas a legenda (título foi removido)
    plt.subplots_adjust(top=0.92)
    
    plt.xticks(rotation=45, ha='right')
    fig.tight_layout(rect=[0, 0, 1, 0.9])  # Ajustar layout para acomodar título e legenda
    
    # Salvar o gráfico em um buffer
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=120, bbox_inches='tight')
    buf.seek(0)
    graficos_buffer.append(buf)
    
    plt.close(fig)
    
    return graficos_buffer

def validar_arquivo_conferencia(arquivo):
    """
    Valida o arquivo de conferência antes do processamento.
    
    Args:
        arquivo: caminho do arquivo CSV de conferência
        
    Returns:
        tuple: (bool, str) - (válido, mensagem de erro)
    """
    try:
        # Tentar ler o arquivo
        df = pd.read_csv(arquivo, dtype=str)
        
        # Validar colunas necessárias
        colunas_necessarias = [
            'AT/TO',
            'AT/TO Validation Status',
            'Total Final Orders Inside AT/TO',
            'Validation Start Time',
            'Validation End Time'
        ]
        
        colunas_faltantes = [col for col in colunas_necessarias if col not in df.columns]
        if colunas_faltantes:
            return False, f"Arquivo não contém as colunas necessárias: {', '.join(colunas_faltantes)}"
            
        # Validar se há dados
        if df.empty:
            return False, "O arquivo está vazio"
            
        # Validar se há rotas validadas
        rotas_validadas = df[df['AT/TO Validation Status'] == 'Validated']
        if rotas_validadas.empty:
            return False, "Não há rotas validadas no arquivo"
            
        # Validar formato das datas
        for col in ['Validation Start Time', 'Validation End Time']:
            try:
                pd.to_datetime(df[col], errors='raise')
            except:
                return False, f"Formato de data inválido na coluna {col}"
                
        # Validar formato dos números
        try:
            pd.to_numeric(df['Total Final Orders Inside AT/TO'], errors='raise')
        except:
            return False, "Formato inválido na coluna de quantidade de pedidos"
            
        return True, "Arquivo válido"
        
    except pd.errors.EmptyDataError:
        return False, "O arquivo está vazio"
    except pd.errors.ParserError:
        return False, "Erro ao processar o arquivo. Verifique se o formato está correto (CSV)"
    except FileNotFoundError:
        return False, "Arquivo não encontrado"
    except Exception as e:
        return False, f"Erro ao validar arquivo: {str(e)}"

def validar_arquivo_expedicao(arquivo):
    """
    Valida o arquivo de expedição antes do processamento.
    
    Args:
        arquivo: caminho do arquivo CSV de expedição
        
    Returns:
        tuple: (bool, str) - (válido, mensagem de erro)
    """
    try:
        # Tentar ler o arquivo
        df = pd.read_csv(arquivo, dtype=str)
        
        # Validar colunas necessárias
        colunas_necessarias = [
            'Task ID',
            'Agency',
            'Driver name',
            'SPX tracking num',
            'Status'  # Adicionada coluna de status como obrigatória
        ]
        
        colunas_faltantes = [col for col in colunas_necessarias if col not in df.columns]
        if colunas_faltantes:
            return False, f"Arquivo não contém as colunas necessárias: {', '.join(colunas_faltantes)}"
            
        # Validar se há dados
        if df.empty:
            return False, "O arquivo está vazio"
            
        return True, "Arquivo válido"
        
    except pd.errors.EmptyDataError:
        return False, "O arquivo está vazio"
    except pd.errors.ParserError:
        return False, "Erro ao processar o arquivo. Verifique se o formato está correto (CSV)"
    except FileNotFoundError:
        return False, "Arquivo não encontrado"
    except Exception as e:
        return False, f"Erro ao validar arquivo: {str(e)}"

def identificar_rotas_outras_janelas(df_assignment, current_window_key):
    """
    Identifica rotas que pertencem a outras janelas de carregamento
    
    Args:
        df_assignment: DataFrame com os dados do arquivo de assignment
        current_window_key: Chave da janela atual (MANHA, TARDE, NOITE)
    
    Returns:
        DataFrame com coluna adicional indicando se a rota é de outra janela
    """
    config_manager = ConfigManager()
    
    # Adicionar coluna para marcar rotas de outras janelas
    df = df_assignment.copy()
    df['janela_carregamento'] = ''
    df['outra_janela'] = False
    
    # Processar cada rota
    for idx, row in df.iterrows():
        delivery_date = row['Delivery Date'] if 'Delivery Date' in row else None
        
        if delivery_date:
            # Converter para formato YYYY-MM-DD
            if isinstance(delivery_date, pd.Timestamp):
                delivery_date = delivery_date.strftime('%Y-%m-%d')
            elif isinstance(delivery_date, str):
                try:
                    # Tentar converter string para data
                    delivery_date = pd.to_datetime(delivery_date).strftime('%Y-%m-%d')
                except:
                    delivery_date = None
        
        if delivery_date:
            window_key, is_other_window = config_manager.get_window_by_delivery_date(
                delivery_date,
                current_window_key
            )
            df.at[idx, 'janela_carregamento'] = config_manager.get_window_name(window_key)
            df.at[idx, 'outra_janela'] = is_other_window
    
    return df

def identificar_rotas_no_piso(df_auditoria, df_expedicao):
    """
    Identifica rotas que estão no piso com base nas seguintes regras:
    1. Rotas que estão na planilha de conferência mas não foram validadas
    2. Rotas que estão validadas na conferência mas estão como "Processing" ou "Processed" no assignment
    
    Args:
        df_auditoria: DataFrame com dados da planilha de conferência
        df_expedicao: DataFrame com dados da planilha de assignment
        
    Returns:
        tuple: (set de rotas no piso, dict com status de cada rota)
            - rotas_no_piso: conjunto com todas as rotas no piso
            - status_rotas: dicionário com status de cada rota (validada ou não)
    """
    rotas_no_piso = set()
    status_rotas = {}  # Para armazenar se cada rota foi validada ou não
    
    # 1. Identificar rotas não validadas na conferência
    rotas_nao_validadas = set(df_auditoria[df_auditoria['AT/TO Validation Status'] != 'Validated']['AT/TO'])
    rotas_no_piso.update(rotas_nao_validadas)
    for rota in rotas_nao_validadas:
        status_rotas[rota] = 'nao_validada'
    
    # 2. Identificar rotas validadas na conferência mas com status "Processing" ou "Processed" no assignment
    rotas_validadas = set(df_auditoria[df_auditoria['AT/TO Validation Status'] == 'Validated']['AT/TO'])
    status_validos = ['Processing', 'Processed']
    
    for rota in rotas_validadas:
        # Verificar se a rota existe no assignment e tem status válido
        if rota in df_expedicao['Task ID'].values:
            status_rota = df_expedicao[df_expedicao['Task ID'] == rota]['Status'].iloc[0]
            if status_rota in status_validos:
                rotas_no_piso.add(rota)
                status_rotas[rota] = 'validada'
    
    return rotas_no_piso, status_rotas

def main(expedicao_files, conferencia_file, output_dir, info_adicional, current_window_key="MANHA"):
    """
    Função principal para gerar os relatórios
    
    Args:
        expedicao_files: lista de caminhos de arquivos CSV de expedição
        conferencia_file: caminho do arquivo CSV de conferência
        output_dir: diretório de saída para os relatórios
        info_adicional: texto de informações adicionais sobre o fechamento da expedição
        current_window_key: chave da janela atual (MANHA, TARDE, NOITE)
    """
    try:
        # Validar arquivo de conferência
        valido, mensagem = validar_arquivo_conferencia(conferencia_file)
        if not valido:
            raise ValueError(mensagem)
            
        # Validar arquivos de expedição
        for exp_file in expedicao_files:
            valido, mensagem = validar_arquivo_expedicao(exp_file)
            if not valido:
                raise ValueError(f"Erro no arquivo {os.path.basename(exp_file)}: {mensagem}")
        
        # Carregar dados
        df_conferencia = pd.read_csv(conferencia_file)
        
        # Lista para armazenar DataFrames de expedição
        dfs_expedicao = []
        
        # Processar cada arquivo de expedição
        for exp_file in expedicao_files:
            df_exp = pd.read_csv(exp_file)
            # Identificar rotas de outras janelas
            df_exp = identificar_rotas_outras_janelas(df_exp, current_window_key)
            dfs_expedicao.append(df_exp)
        
        # Combinar todos os DataFrames de expedição
        if len(dfs_expedicao) > 1:
            df_expedicao = pd.concat(dfs_expedicao, ignore_index=True)
        else:
            df_expedicao = dfs_expedicao[0]
        
        # 3. Ler arquivo de conferência (já validado)
        df_auditoria = pd.read_csv(conferencia_file, dtype=str)

        # 4. Conversão de datas
        for col in ['Create Time', 'Complete time', 'Driver Assigned Time', 'Agency Assigned Time']:
            if col in df_expedicao.columns:
                df_expedicao[col] = pd.to_datetime(df_expedicao[col], errors='coerce')
        for col in ['Validation Start Time', 'Validation End Time']:
            if col in df_auditoria.columns:
                df_auditoria[col] = pd.to_datetime(df_auditoria[col], errors='coerce')

        # Identificar rotas no piso usando a nova lógica
        rotas_no_piso_set, status_rotas = identificar_rotas_no_piso(df_auditoria, df_expedicao)
        
        # Filtrar apenas rotas validadas (que não estão no piso)
        df_validadas = df_auditoria[
            (df_auditoria['AT/TO Validation Status'] == 'Validated') & 
            (~df_auditoria['AT/TO'].isin(rotas_no_piso_set))
        ].copy()

        # 1. Quantidade de rotas expedidas (validadas e não no piso)
        rotas_expedidas = df_validadas['AT/TO'].nunique()

        # 2. Quantidade de rotas expedidas por transportadora (validadas e não no piso)
        # Garantir 1 linha por AT/TO validado, buscar transportadora e somar pedidos expedidos corretamente
        rotas_validas = df_validadas[['AT/TO', 'Total Final Orders Inside AT/TO']].drop_duplicates(subset=['AT/TO'])
        
        # Buscar informações de transportadora do arquivo de expedição apenas para complementar
        info_transportadora = df_expedicao[['Task ID', 'Agency']].drop_duplicates(subset=['Task ID'])
        
        # Merge com left join para manter todas as rotas validadas
        rotas_validas = rotas_validas.merge(
            info_transportadora,
            left_on='AT/TO',
            right_on='Task ID',
            how='left'
        )
        
        # Tratar casos onde a transportadora não foi encontrada
        rotas_validas['Agency'] = rotas_validas['Agency'].fillna('Não informado')
        
        rotas_por_agencia = rotas_validas.groupby('Agency').agg({
            'AT/TO': 'nunique',
            'Total Final Orders Inside AT/TO': lambda x: x.astype(float).sum()
        }).reset_index()
        
        rotas_por_agencia.columns = ['Transportadora', 'Rotas Expedidas', 'Pedidos Expedidos']
        rotas_por_agencia['Pedidos Expedidos'] = rotas_por_agencia['Pedidos Expedidos'].astype(int)
        
        # Ordenar por número de rotas (decrescente)
        rotas_por_agencia = rotas_por_agencia.sort_values('Rotas Expedidas', ascending=False)

        # Adicionar totais em Rotas por Transportadora
        if not rotas_por_agencia.empty:
            total_rotas = rotas_por_agencia['Rotas Expedidas'].sum()
            total_pedidos = rotas_por_agencia['Pedidos Expedidos'].sum()
            rotas_por_agencia = pd.concat([
                rotas_por_agencia,
                pd.DataFrame([{
                    'Transportadora': 'Total',
                    'Rotas Expedidas': total_rotas,
                    'Pedidos Expedidos': total_pedidos
                }])
            ], ignore_index=True)

        # 3. Horário de início/fim da expedição (validadas)
        inicio_expedicao = df_validadas['Validation Start Time'].min()
        fim_expedicao = df_validadas['Validation End Time'].max()

        # 4. Média de tempo de conferência de uma rota (validadas)
        df_validadas['duracao_conferencia'] = df_validadas['Validation End Time'] - df_validadas['Validation Start Time']
        media_conferencia = df_validadas['duracao_conferencia'].mean()
        media_conferencia_fmt = format_hms(media_conferencia)

        # Nova funcionalidade: Contagem de rotas por tipo de veículo
        # Procurar por uma coluna que represente o tipo de veículo
        vehicle_type_columns = [col for col in df_expedicao.columns 
                               if any(term in col.lower() for term in 
                                     ['vehicle', 'veic', 'veíc', 'car', 'carro', 'truck', 'caminhão', 'caminhao', 
                                      'tipo', 'type', 'model', 'modelo', 'transporte', 'transport'])]
        
        # Preferência para colunas com 'tipo' + 'veículo' ou 'vehicle' + 'type'
        specific_columns = [col for col in vehicle_type_columns 
                           if ('tipo' in col.lower() and ('veic' in col.lower() or 'veíc' in col.lower())) or 
                              ('vehicle' in col.lower() and 'type' in col.lower())]
        
        vehicle_type_column = next(iter(specific_columns), None) if specific_columns else next(iter(vehicle_type_columns), None)
        
        # Se ainda não encontrou a coluna, use o nome mais comum
        if not vehicle_type_column:
            common_names = ['Vehicle Type', 'Tipo de Veículo', 'Tipo de Veiculo', 'Car Type', 'Tipo', 'Type']
            for name in common_names:
                if name in df_expedicao.columns:
                    vehicle_type_column = name
                    break
        
        if vehicle_type_column in df_expedicao.columns:
            # Contar rotas expedidas por tipo de veículo
            # Usar as rotas validadas como base e buscar informação de veículo do arquivo de expedição
            info_veiculo = df_expedicao[['Task ID', vehicle_type_column]].drop_duplicates(subset=['Task ID'])
            
            rotas_validas_com_veiculo = rotas_validas.merge(
                info_veiculo,
                left_on='AT/TO',
                right_on='Task ID', 
                how='left'
            )
            
            # Substituir valores nulos ou vazios por "Não informado"
            rotas_validas_com_veiculo[vehicle_type_column] = rotas_validas_com_veiculo[vehicle_type_column].fillna('Não informado')
            rotas_validas_com_veiculo.loc[rotas_validas_com_veiculo[vehicle_type_column].str.strip() == '', vehicle_type_column] = 'Não informado'
            
            # Normalizar/agrupar tipos de veículos conforme regras de negócio
            def normalizar_tipo_veiculo(tipo):
                if pd.isna(tipo) or str(tipo).strip() == '':
                    return 'Não informado'
                
                tipo = str(tipo).upper()
                
                # Regras de agrupamento
                if 'PASSEIO' in tipo:
                    return 'PASSEIO'
                elif 'FIORINO' in tipo:
                    return 'FIORINO'
                elif 'MOTO' in tipo:
                    return 'MOTO'
                elif 'VAN' in tipo:
                    return 'VAN'
                else:
                    return tipo
            
            # Aplicar normalização
            rotas_validas_com_veiculo['Tipo_Normalizado'] = rotas_validas_com_veiculo[vehicle_type_column].apply(normalizar_tipo_veiculo)
            
            rotas_por_veiculo = rotas_validas_com_veiculo.groupby('Tipo_Normalizado').agg({
                'AT/TO': 'nunique'
            }).reset_index()
            
            rotas_por_veiculo.columns = ['Tipo de Veículo', 'Rotas Expedidas']
            
            # Ordenar por número de rotas (decrescente)
            rotas_por_veiculo = rotas_por_veiculo.sort_values('Rotas Expedidas', ascending=False)
            
            # Adicionar totais em Rotas por Tipo de Veículo
            if not rotas_por_veiculo.empty:
                total_rotas = rotas_por_veiculo['Rotas Expedidas'].sum()
                rotas_por_veiculo = pd.concat([
                    rotas_por_veiculo,
                    pd.DataFrame([{
                        'Tipo de Veículo': 'Total',
                        'Rotas Expedidas': total_rotas
                    }])
                ], ignore_index=True)
        else:
            # Criar dataframe vazio se o tipo de veículo não estiver disponível
            rotas_por_veiculo = pd.DataFrame(columns=['Tipo de Veículo', 'Rotas Expedidas'])
            # Adicionar uma linha indicando que o tipo de veículo não está disponível
            rotas_por_veiculo = pd.DataFrame([{
                'Tipo de Veículo': 'Informação não disponível',
                'Rotas Expedidas': 0
            }, {
                'Tipo de Veículo': 'Total',
                'Rotas Expedidas': 0
            }])
            print(f"Aviso: Coluna de tipo de veículo não encontrada nos dados de expedição. Colunas disponíveis: {', '.join(df_expedicao.columns)}")

        # 5. Métricas por operador: tempo médio de conferência e tempo médio de ociosidade
        metricas_operador = []
        def extrair_nome_operador(op):
            # Remove prefixo [opsXXXXX] se existir e aplica title case
            return re.sub(r'^\[.*?\]', '', op).strip().title()
        if 'Validation Operator' in df_validadas.columns and 'Validation Start Time' in df_validadas.columns and 'Validation End Time' in df_validadas.columns:
            for operador, grupo in df_validadas.groupby('Validation Operator'):
                grupo = grupo.sort_values('Validation Start Time')
                # Tempo médio de conferência
                duracoes = grupo['Validation End Time'] - grupo['Validation Start Time']
                media_conferencia = duracoes.mean()
                # Tempo médio de ociosidade
                ociosidade = grupo['Validation Start Time'].shift(-1) - grupo['Validation End Time']
                media_ociosidade = ociosidade[:-1].mean()  # ignora o último, que não tem próximo
                metricas_operador.append({
                    'Operador': extrair_nome_operador(operador),
                    'Conferência': format_hms(media_conferencia),
                    'Ociosidade': format_hms(media_ociosidade)
                })
        metricas_operador_df = pd.DataFrame(metricas_operador)

        # Adicionar linha de médias finais
        if not metricas_operador_df.empty:
            conferencia_td = pd.to_timedelta(metricas_operador_df['Conferência'])
            ociosidade_td = pd.to_timedelta(metricas_operador_df['Ociosidade'])
            media_geral_conferencia = conferencia_td.mean()
            media_geral_ociosidade = ociosidade_td.mean()
            metricas_operador_df = pd.concat([
                metricas_operador_df,
                pd.DataFrame([{
                    'Operador': 'Média Geral',
                    'Conferência': format_hms(media_geral_conferencia),
                    'Ociosidade': format_hms(media_geral_ociosidade)
                }])
            ], ignore_index=True)
            # Ordenar por tempo de conferência decrescente, mantendo 'Média Geral' por último
            media_geral = metricas_operador_df[metricas_operador_df['Operador'] == 'Média Geral']
            outros = metricas_operador_df[metricas_operador_df['Operador'] != 'Média Geral'].copy()
            outros['Conferência_td'] = pd.to_timedelta(outros['Conferência'])
            outros = outros.sort_values('Conferência_td', ascending=False).drop(columns='Conferência_td')
            metricas_operador_df = pd.concat([outros, media_geral], ignore_index=True)

        # 6. Rotas e pedidos expedidos por hora (validadas)
        df_validadas['hora_expedicao'] = df_validadas['Validation End Time'].dt.floor('H')
        rotas_por_hora = df_validadas.groupby('hora_expedicao')['AT/TO'].nunique().reset_index()
        rotas_por_hora.columns = ['Hora', 'Rotas Expedidas']
        pedidos_por_hora = df_validadas.groupby('hora_expedicao')['Total Final Orders Inside AT/TO'].apply(lambda x: x.astype(float).sum()).reset_index()
        pedidos_por_hora.columns = ['Hora', 'Pedidos Expedidos']
        expedidos_por_hora = pd.merge(rotas_por_hora, pedidos_por_hora, on='Hora')
        expedidos_por_hora['Pedidos Expedidos'] = expedidos_por_hora['Pedidos Expedidos'].astype(int)

        # Adicionar totais em Rotas e Pedidos Expedidos por Hora
        if not expedidos_por_hora.empty:
            total_rotas = expedidos_por_hora['Rotas Expedidas'].sum()
            total_pedidos = expedidos_por_hora['Pedidos Expedidos'].sum()
            expedidos_por_hora = pd.concat([
                expedidos_por_hora,
                pd.DataFrame([{
                    'Hora': 'Total',
                    'Rotas Expedidas': total_rotas,
                    'Pedidos Expedidos': total_pedidos
                }])
            ], ignore_index=True)

        def hora_intervalo_str(h):
            if pd.isnull(h):
                return ''
            if isinstance(h, pd.Timestamp):
                hora_ini = h.strftime('%H:%M')
                hora_fim = (h + pd.Timedelta(hours=1)).strftime('%H:%M')
                return f'{hora_ini} às {hora_fim}'
            return str(h)

        expedidos_por_hora['Hora'] = expedidos_por_hora['Hora'].apply(hora_intervalo_str)

        # Gerar gráficos de rotas e pedidos por hora
        graficos_buffer = criar_grafico_hora_a_hora(expedidos_por_hora)

        # 7. Rotas não conferidas (no piso)
        rotas_conferidas = set(df_validadas['AT/TO'].unique())
        
        # Filtrar apenas rotas com status "Processing" ou "Processed" do arquivo de expedição
        status_validos = ['Processing', 'Processed']
        df_expedicao_filtrado = df_expedicao[df_expedicao['Status'].isin(status_validos)]
        rotas_assignment = set(df_expedicao_filtrado['Task ID'].unique())
        
        # Encontrar todas as rotas que não foram validadas no arquivo de conferência
        todas_rotas_auditoria = set(df_auditoria['AT/TO'].unique())
        rotas_nao_validadas = todas_rotas_auditoria - rotas_conferidas
        
        # Unir rotas do assignment (apenas com status válido) que não foram conferidas 
        # com rotas da auditoria que não foram validadas
        rotas_nao_conferidas = (rotas_assignment - rotas_conferidas) | rotas_nao_validadas
        
        # Atualizar a criação do DataFrame de rotas não conferidas
        pacotes_por_rota = []
        for rota in rotas_no_piso_set:
            # Buscar informações da rota no arquivo de expedição
            info_rota_exp = df_expedicao[df_expedicao['Task ID'] == rota]
            info_rota_aud = df_auditoria[df_auditoria['AT/TO'] == rota]
            
            if not info_rota_exp.empty:
                transportadora = info_rota_exp.iloc[0].get('Agency', 'Não atribuído')
                motorista = info_rota_exp.iloc[0].get('Driver name', 'Não atribuído')
            else:
                transportadora = 'Não atribuído'
                motorista = 'Não atribuído'
            
            # Definir a quantidade de pacotes com base no status da rota
            if not info_rota_aud.empty:
                if status_rotas.get(rota) == 'validada':
                    # Se a rota foi validada, usar o total final
                    try:
                        pacotes = int(float(info_rota_aud.iloc[0].get('Total Final Orders Inside AT/TO', 0)))
                    except (ValueError, TypeError):
                        pacotes = 0
                else:
                    # Se a rota não foi validada, usar o total inicial
                    try:
                        pacotes = int(float(info_rota_aud.iloc[0].get('Total Initial Orders Inside AT/TO', 0)))
                    except (ValueError, TypeError):
                        pacotes = 0
            else:
                # Se não encontrar na auditoria, tentar no arquivo de expedição
                pacotes = len(info_rota_exp['SPX tracking num'].unique()) if not info_rota_exp.empty else 0
            
            pacotes_por_rota.append({
                'Rota': rota,
                'Transportadora': transportadora,
                'Motorista': motorista,
                'Pacotes': pacotes
            })
        
        # Criar DataFrame com as rotas não conferidas
        pacotes_por_rota_df = pd.DataFrame(pacotes_por_rota)
        
        # Calcular totais
        total_rotas = len(rotas_no_piso_set)
        total_pacotes = pacotes_por_rota_df['Pacotes'].sum() if not pacotes_por_rota_df.empty else 0
        
        # Criar a linha de totais
        totais_df = pd.DataFrame([{
            'Rota': f'Total ({total_rotas} rotas)',
            'Transportadora': '',
            'Motorista': '',
            'Pacotes': total_pacotes
        }])
        
        rotas_nao_conferidas_df = pd.concat([pacotes_por_rota_df, totais_df], ignore_index=True) if not pacotes_por_rota_df.empty else totais_df
        
        # Substituir valores nulos/nan por 'Não atribuído' nas colunas relevantes (exceto linha de total)
        if not rotas_nao_conferidas_df.empty:
            total_row_index = rotas_nao_conferidas_df[rotas_nao_conferidas_df['Rota'].astype(str).str.startswith('Total')].index
            non_total_rows = ~rotas_nao_conferidas_df.index.isin(total_row_index)
            
            # Aplicar substituições apenas nas linhas que não são totais
            for col in ['Transportadora', 'Motorista']:
                if col in rotas_nao_conferidas_df.columns:
                    # Substituir NaN
                    rotas_nao_conferidas_df.loc[non_total_rows, col] = rotas_nao_conferidas_df.loc[non_total_rows, col].fillna('Não atribuído')
                    
                    # Substituir strings vazias
                    mask_empty = (rotas_nao_conferidas_df[col].astype(str).str.strip() == '') & non_total_rows
                    rotas_nao_conferidas_df.loc[mask_empty, col] = 'Não atribuído'
                    
                    # Substituir o texto literal 'nan' (case insensitive)
                    mask_nan = (rotas_nao_conferidas_df[col].astype(str).str.lower() == 'nan') & non_total_rows
                    rotas_nao_conferidas_df.loc[mask_nan, col] = 'Não atribuído'
        
        # Aplicar title case apenas na coluna de Motorista (mantendo o anterior)
        if 'Motorista' in rotas_nao_conferidas_df.columns:
            rotas_nao_conferidas_df['Motorista'] = rotas_nao_conferidas_df['Motorista'].astype(str).str.title()
            # Garantir que "nan" depois de title case não seja exibido como "Nan"
            rotas_nao_conferidas_df.loc[rotas_nao_conferidas_df['Motorista'] == 'Nan', 'Motorista'] = 'Não atribuído'

        # Montar DataFrame final para exportação
        resumo = []
        
        # Usar primariamente o arquivo de auditoria para as contagens
        rotas_programadas = df_auditoria['AT/TO'].nunique()
        pedidos_programados = df_auditoria['Total Initial Orders Inside AT/TO'].astype(float).sum()
        
        # Rotas expedidas e pedidos expedidos vêm das rotas validadas
        rotas_expedidas = df_validadas['AT/TO'].nunique()
        pedidos_expedidos = df_validadas['Total Final Orders Inside AT/TO'].astype(float).sum()
        
        # Rotas não conferidas são as que estão no arquivo de auditoria mas não foram validadas
        rotas_auditoria = set(df_auditoria['AT/TO'].unique())
        rotas_validadas = set(df_validadas['AT/TO'].unique())
        rotas_nao_conferidas = rotas_auditoria - rotas_validadas
        rotas_no_piso = len(rotas_nao_conferidas)
        
        resumo.append({'Métrica': 'Quantidade de rotas programadas', 'Valor': rotas_programadas})
        resumo.append({'Métrica': 'Quantidade de rotas expedidas', 'Valor': rotas_expedidas})
        resumo.append({'Métrica': 'Quantidade de pedidos programados', 'Valor': int(pedidos_programados)})
        resumo.append({'Métrica': 'Quantidade de pedidos expedidos', 'Valor': int(pedidos_expedidos)})
        resumo.append({'Métrica': 'Quantidade de rotas que ficaram no piso', 'Valor': rotas_no_piso})
        resumo.append({'Métrica': 'Horário de início da expedição', 'Valor': format_hms(inicio_expedicao)})
        resumo.append({'Métrica': 'Horário de fim da expedição', 'Valor': format_hms(fim_expedicao)})
        resumo.append({'Métrica': 'Média de tempo de giro de bancada', 'Valor': media_conferencia_fmt})
        
        resumo_df = pd.DataFrame(resumo)

        # Exportar todos os resultados para um único CSV, separados por seções
        csv_path = os.path.join(output_dir, 'resumo_expedicao.csv')
        with open(csv_path, 'w', encoding='utf-8', newline='') as f:
            resumo_df.to_csv(f, index=False)
            f.write('\n')
            f.write('Rotas expedidas por transportadora\n')
            rotas_por_agencia.to_csv(f, index=False)
            f.write('\n')
            f.write('Rotas expedidas por tipo de veículo\n')
            rotas_por_veiculo.to_csv(f, index=False)
            f.write('\n')
            f.write('Rotas e pedidos expedidos por hora\n')
            expedidos_por_hora.to_csv(f, index=False)
            f.write('\n')
            f.write('Média de tempo entre uma conferência e outra por usuário\n')
            metricas_operador_df.to_csv(f, index=False)
            f.write('\n')
            f.write('Rotas NS - ficaram no piso\n')
            rotas_nao_conferidas_df.to_csv(f, index=False)
            if info_adicional.strip():
                f.write('\nInformações adicionais sobre o fechamento da expedição:\n')
                f.write(info_adicional.strip() + '\n')

        print(f'Arquivo {csv_path} gerado com sucesso!')

        # Função para converter DataFrame em dados para Table do ReportLab
        def df_to_tabledata(df):
            data = [list(df.columns)] + df.astype(str).values.tolist()
            return data

        # Cores Shopee
        SHOPEE_ORANGE = colors.HexColor('#FF5722')
        SHOPEE_BG = colors.HexColor('#F5F5F5')
        SHOPEE_TEXT = colors.HexColor('#222222')

        # Criar PDF
        pdf_path = os.path.join(output_dir, 'relatorio_expedicao.pdf')
        doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        elements = []

        styles = getSampleStyleSheet()
        style_title = ParagraphStyle(
            'ShopeeTitle',
            parent=styles['Heading1'],
            fontSize=22,
            textColor=SHOPEE_ORANGE,
            spaceAfter=18,
            alignment=TA_CENTER  # Centraliza o texto
        )
        style_section = ParagraphStyle(
            'ShopeeSection',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=SHOPEE_ORANGE,
            spaceAfter=12,
            alignment=TA_CENTER  # Centraliza o título da seção
        )
        style_normal = ParagraphStyle('ShopeeNormal', parent=styles['Normal'], fontSize=10, textColor=SHOPEE_TEXT)

        # Título principal
        elements.append(Paragraph('Relatório de Expedição - Shopee', style_title))
        
        # Adicionar turno selecionado
        config_manager = ConfigManager()
        config = config_manager.load_config()
        window_config = config.get(current_window_key, {})
        if window_config:
            window_display = config_manager.format_window_display(current_window_key, window_config)
            elements.append(Paragraph(window_display, style_section))
        
        elements.append(Spacer(1, 18))

        # Seção: Métricas Gerais
        table = Table(df_to_tabledata(resumo_df), hAlign='CENTER')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), SHOPEE_ORANGE),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 11),
            ('FONTSIZE', (0,1), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, SHOPEE_BG]),
            ('GRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
        ]))
        elements.append(KeepTogether([
            Paragraph('Métricas Gerais', style_section),
            table,
            Spacer(1, 18)
        ]))

        # Seção: Rotas Expedidas por Transportadora
        table = Table(df_to_tabledata(rotas_por_agencia), hAlign='CENTER')
        
        # Determinar a linha de totais (normalmente é a última linha)
        total_row = len(df_to_tabledata(rotas_por_agencia)) - 1
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), SHOPEE_ORANGE),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 11),
            ('FONTSIZE', (0,1), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, SHOPEE_BG]),
            ('GRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
            # Destacar a linha de totais
            ('BACKGROUND', (0,total_row), (-1,total_row), colors.HexColor('#F5F5F5')),
            ('FONTNAME', (0,total_row), (-1,total_row), 'Helvetica-Bold'),
            ('LINEABOVE', (0,total_row), (-1,total_row), 1, colors.HexColor('#BBBBBB')),
            ('LINEBELOW', (0,total_row), (-1,total_row), 1, colors.HexColor('#BBBBBB')),
        ]))
        elements.append(KeepTogether([
            Paragraph('Rotas Expedidas por Transportadora', style_section),
            table,
            Spacer(1, 18)
        ]))

        # Seção: Rotas Expedidas por Tipo de Veículo
        table = Table(df_to_tabledata(rotas_por_veiculo), hAlign='CENTER')
        
        # Determinar a linha de totais (normalmente é a última linha)
        total_row = len(df_to_tabledata(rotas_por_veiculo)) - 1
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), SHOPEE_ORANGE),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 11),
            ('FONTSIZE', (0,1), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, SHOPEE_BG]),
            ('GRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
            # Destacar a linha de totais
            ('BACKGROUND', (0,total_row), (-1,total_row), colors.HexColor('#F5F5F5')),
            ('FONTNAME', (0,total_row), (-1,total_row), 'Helvetica-Bold'),
            ('LINEABOVE', (0,total_row), (-1,total_row), 1, colors.HexColor('#BBBBBB')),
            ('LINEBELOW', (0,total_row), (-1,total_row), 1, colors.HexColor('#BBBBBB')),
        ]))
        elements.append(KeepTogether([
            Paragraph('Rotas Expedidas por Tipo de Veículo', style_section),
            table,
            Spacer(1, 18)
        ]))

        # Seção: Rotas e Pedidos Expedidos por Hora
        # Usar PageBreak para garantir que esta seção comece em uma nova página
        elements.append(PageBreak())
        
        # Título da seção
        section_title = Paragraph('Rotas e Pedidos Expedidos por Hora', style_section)
        
        table = Table(df_to_tabledata(expedidos_por_hora), hAlign='CENTER')
        
        # Determinar a linha de totais (normalmente é a última linha)
        total_row = len(df_to_tabledata(expedidos_por_hora)) - 1
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), SHOPEE_ORANGE),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 11),
            ('FONTSIZE', (0,1), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, SHOPEE_BG]),
            ('GRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
            # Destacar a linha de totais
            ('BACKGROUND', (0,total_row), (-1,total_row), colors.HexColor('#F5F5F5')),
            ('FONTNAME', (0,total_row), (-1,total_row), 'Helvetica-Bold'),
            ('LINEABOVE', (0,total_row), (-1,total_row), 1, colors.HexColor('#BBBBBB')),
            ('LINEBELOW', (0,total_row), (-1,total_row), 1, colors.HexColor('#BBBBBB')),
        ]))
        
        # Primeiro conteúdo da página: título e tabela (mantendo juntos)
        elements.append(KeepTogether([
            section_title,
            Spacer(1, 10),
            table,
            Spacer(1, 20)
        ]))
        
        # Adicionar gráficos na mesma página
        if graficos_buffer and len(graficos_buffer) > 0:
            # Adicionar gráfico logo após a tabela
            img_width = 18 * cm
            img_height = 9 * cm
            elements.append(Image(graficos_buffer[0], width=img_width, height=img_height))
            elements.append(Spacer(1, 18))

        # Seção: Tempo médio de conferência e ociosidade por operador
        table = Table(df_to_tabledata(metricas_operador_df), hAlign='CENTER')
        
        # Determinar a linha de média geral (normalmente é a última linha)
        total_row = len(df_to_tabledata(metricas_operador_df)) - 1
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), SHOPEE_ORANGE),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 11),
            ('FONTSIZE', (0,1), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, SHOPEE_BG]),
            ('GRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
            # Destacar a linha de média geral
            ('BACKGROUND', (0,total_row), (-1,total_row), colors.HexColor('#F5F5F5')),
            ('FONTNAME', (0,total_row), (-1,total_row), 'Helvetica-Bold'),
            ('LINEABOVE', (0,total_row), (-1,total_row), 1, colors.HexColor('#BBBBBB')),
            ('LINEBELOW', (0,total_row), (-1,total_row), 1, colors.HexColor('#BBBBBB')),
        ]))
        elements.append(KeepTogether([
            Paragraph('Tempo Médio de Conferência e Ociosidade por Operador', style_section),
            table,
            Spacer(1, 18)
        ]))

        # Seção: Rotas NS - Ficaram no Piso
        table = Table(df_to_tabledata(rotas_nao_conferidas_df), hAlign='CENTER')
        
        # Determinar a linha de totais (normalmente é a última linha)
        total_row = len(df_to_tabledata(rotas_nao_conferidas_df)) - 1
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), SHOPEE_ORANGE),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 11),
            ('FONTSIZE', (0,1), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, SHOPEE_BG]),
            ('GRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
            # Destacar a linha de totais
            ('BACKGROUND', (0,total_row), (-1,total_row), colors.HexColor('#F5F5F5')),
            ('FONTNAME', (0,total_row), (-1,total_row), 'Helvetica-Bold'),
            ('LINEABOVE', (0,total_row), (-1,total_row), 1, colors.HexColor('#BBBBBB')),
            ('LINEBELOW', (0,total_row), (-1,total_row), 1, colors.HexColor('#BBBBBB')),
        ]))
        elements.append(KeepTogether([
            Paragraph('Rotas NS - Ficaram no Piso', style_section),
            table,
            Spacer(1, 18)
        ]))

        # Adicionar informações adicionais ao PDF
        if info_adicional.strip():
            elements.append(PageBreak())
            elements.append(Paragraph('Informações adicionais sobre o fechamento da expedição:', style_section))
            elements.append(Paragraph(info_adicional.strip().replace('\n', '<br/>'), style_normal))

        # Gerar PDF
        print('Gerando PDF...')
        doc.build(elements)
        print(f'Relatório PDF gerado com sucesso: {pdf_path}')

        # Geração do arquivo Excel estilizado
        xlsx_path = os.path.join(output_dir, 'relatorio_expedicao.xlsx')
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            resumo_df.to_excel(writer, sheet_name='Resumo', index=False, startrow=2)
            rotas_por_agencia.to_excel(writer, sheet_name='Rotas por Transportadora', index=False, startrow=2)
            rotas_por_veiculo.to_excel(writer, sheet_name='Rotas por Tipo de Veículo', index=False, startrow=2)
            expedidos_por_hora.to_excel(writer, sheet_name='Hora a Hora', index=False, startrow=2)
            metricas_operador_df.to_excel(writer, sheet_name='Métricas por Operador', index=False, startrow=2)
            rotas_nao_conferidas_df.to_excel(writer, sheet_name='Rotas NS', index=False, startrow=2)
            
            if info_adicional.strip():
                info_df = pd.DataFrame({'Informações adicionais sobre o fechamento da expedição:': [info_adicional.strip()]})
                info_df.to_excel(writer, sheet_name='Informações Adicionais', index=False)

            wb = writer.book
            orange_fill = PatternFill(start_color='FF5722', end_color='FF5722', fill_type='solid')
            white_font = Font(color='FFFFFF', bold=True)
            left_align = Alignment(horizontal='left', vertical='center')
            border = Border(left=Side(style='thin', color='DDDDDD'),
                            right=Side(style='thin', color='DDDDDD'),
                            top=Side(style='thin', color='DDDDDD'),
                            bottom=Side(style='thin', color='DDDDDD'))
            title_font = Font(color='FF5722', bold=True, size=16)

            titulos = {
                'Resumo': 'Métricas Gerais',
                'Rotas por Transportadora': 'Rotas Expedidas por Transportadora',
                'Rotas por Tipo de Veículo': 'Rotas Expedidas por Tipo de Veículo',
                'Hora a Hora': 'Rotas e Pedidos Expedidos por Hora',
                'Métricas por Operador': 'Tempo Médio de Conferência e Ociosidade por Operador',
                'Rotas NS': 'Rotas NS - Ficaram no Piso'
            }

            total_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
            bold_font = Font(bold=True)
            thick_border = Border(top=Side(style='medium', color='BBBBBB'),
                                 bottom=Side(style='medium', color='BBBBBB'),
                                 left=Side(style='thin', color='DDDDDD'),
                                 right=Side(style='thin', color='DDDDDD'))
            
            for sheet_name in ['Resumo', 'Rotas por Transportadora', 'Rotas por Tipo de Veículo', 'Hora a Hora', 'Métricas por Operador', 'Rotas NS']:
                if sheet_name in wb.sheetnames:  # Verificar se a aba existe
                    ws = wb[sheet_name]
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
                    cell = ws.cell(row=1, column=1)
                    cell.value = titulos[sheet_name]
                    cell.font = title_font
                    cell.alignment = left_align
                    
                    # Estilizar cabeçalho
                    for cell in ws[3]:
                        cell.fill = orange_fill
                        cell.font = white_font
                        cell.alignment = left_align
                        cell.border = border
                    
                    # Estilizar linhas de dados
                    max_row = ws.max_row
                    for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_row=max_row), 4):
                        # Verificar se é a última linha (totais)
                        is_total_row = False
                        first_cell_value = ws.cell(row=row_idx, column=1).value
                        if first_cell_value and (str(first_cell_value).lower().startswith('total') or 
                                                str(first_cell_value).lower() == 'média geral'):
                            is_total_row = True
                        
                        for cell in row:
                            cell.alignment = left_align
                            
                            if is_total_row:
                                # Aplicar estilo especial para linha de totais
                                cell.fill = total_fill
                                cell.font = bold_font
                                cell.border = thick_border
                            else:
                                # Estilo normal para outras linhas
                                cell.border = border
                    for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
                        max_length = 0
                        for cell in col:
                            if cell.value is not None:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                        col_letter = get_column_letter(idx + 1)
                        ws.column_dimensions[col_letter].width = max_length + 2

        print(f'Arquivo {xlsx_path} gerado com sucesso!')
    except Exception as e:
        print(f"Erro ao processar o relatório: {str(e)}")